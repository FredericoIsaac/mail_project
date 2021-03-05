import pandas as pd
from string import ascii_uppercase as abc
from openpyxl import load_workbook
import win32api


class ExcelMachine:

    def __init__(self, excel_file, sheet=None):
        self.excel_file = excel_file
        self.sheet = sheet if sheet else 0
        self.df = pd.read_excel(self.excel_file, sheet_name=self.sheet)
        self.clients_send_mail = self.to_send_mail()
        self.client_info = self.get_info_clients()

    def to_send_mail(self, column: str = "Enviar Mail"):
        """
        Get the column the true values that are the customers to send mail
        :param column: The column to filter
        :return: Series that has column True
        """
        send_mail = self.df[self.df[column]]  # self.df[self.df[column] == True]
        return send_mail

    def get_info_clients(self):
        """
        Return a Dictionary with info of the company's to send mails
        :return:{
                10100 : {
                     0 :{
                        "Ativo": boolean,
                        "Nº Emp.": str,
                        "EMPRESAS": str,
                        "NIF's": str,
                        "Responsável": str,
                        "Ficheiro" str:
                        "Mail - To": str,
                        "Mail - CC": str,
                        "Mail de Confirmação": boolean,
                        "Observações": str,
                        "Mail Enviado": str,
                        "Submetido": boolean,
                        "Enviar Mail": boolean,
                    }
                    1: {
                        ...: ...,
                    }
                }
        }
        """
        mails_clients_to_dict = self.clients_send_mail.to_dict('records')
        clients_info_dict = {}

        for index in range(len(mails_clients_to_dict)):
            principal_key = mails_clients_to_dict[index]["Nº Emp."]  # 10120
            value = mails_clients_to_dict[index]
            if principal_key in clients_info_dict:
                numb_store = max(clients_info_dict[principal_key].keys())
                key = numb_store + 1
                clients_info_dict[principal_key].update({key: value})
            else:
                clients_info_dict[principal_key] = {0: value}

        return clients_info_dict

    def introduce_info(self, company_number: int, store_number: int, info_saft: str):
        """
        Introduce information into the excel dependent of the param
        :param store_number: Get from get_info_client {10101: {0: {'Ativo': True,...}, 1: {'Ativo': True,...}
        :param company_number: The number of the company in question 10100
        :param info_saft: mail_enviado or mail_erro or saft_submetido
        """
        respective_column = "Mail Enviado"
        input_cell = "true"

        if info_saft == "saft_submetido":
            respective_column = "Submetido"
        elif info_saft == "mail_erro":
            input_cell = "false"

        row = self.df.index[self.df["Nº Emp."] == company_number][0]
        row = row + store_number + 2

        columns_dict = dict(zip(self.df.columns, abc))
        column = columns_dict[respective_column]

        wb = load_workbook(filename=self.excel_file)
        sheets = wb.sheetnames
        if self.sheet == 0:
            sheet = 0
        else:
            sheet = sheets.index(self.sheet)

        print(sheet)
        ws = wb.worksheets[sheet]
        ws[f"{column}{row}"].value = input_cell

        while True:
            try:
                wb.save(filename=self.excel_file)
            except PermissionError:
                win32api.MessageBox(0, "To continue close excel file and click ok.", "Excel File Open")
            else:
                break

ALERT = """
Verificar a entrada de informação em 
excel.introduce_info(10101, 0, "mail_enviado")
supostament no main esta a informação toda da empresa com diferentes lojas
tenho que ver como vou processar a informação no main para perceber como vou introduzir os parametros:
:param store_number: Get from get_info_client {10101: {0: {'Ativo': True,...}, 1: {'Ativo': True,...}
:param company_number: The number of the company in question 10100
"""
